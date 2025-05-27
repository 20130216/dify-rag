# File: i_excel_md_convert.py

import sys
import pandas as pd
import os
import re
import datetime
import zipfile # 用于处理openpyxl可能遇到的压缩文件错误

try:
    import xlrd # 用于处理 .xls 文件
except ImportError:
    print("未找到 xlrd 库。要处理 .xls 文件，请安装它 (例如: `pip install xlrd==1.2.0`)")

try:
    import openpyxl # 用于处理 .xlsx 文件
    from openpyxl.utils.exceptions import InvalidFileException
    from openpyxl.cell import Cell as OpenpyxlCell # 导入 Cell 类型用于类型检查
except ImportError:
    print("未找到 openpyxl 库。要处理 .xlsx 文件，请安装它 (例如: `pip install openpyxl`)")

# 尝试从 dify_rag 导入 Document。如果未找到，则定义一个简单的模拟类。
try:
    from dify_rag.models.document import Document
except ImportError:
    print("警告: 未找到 `dify_rag.models.document.Document`。正在使用模拟 Document 类。")
    class Document:
        def __init__(self, page_content: str, metadata: dict):
            self.page_content = page_content
            self.metadata = metadata
        def __repr__(self):
            return f"Document(page_content='{self.page_content[:50]}...', metadata={self.metadata})"

# --- 辅助函数 (改编自 excel_md_convert.py) ---

def _df_to_md(df: pd.DataFrame, table_style: str = "simple") -> str:
    """
    将 DataFrame 转换为 Markdown 表格字符串。
    table_style: "simple" 表示基本 Markdown 表格 (默认), "fancy" 表示更高级的格式。
    注意：此函数会处理列名，将 "UnnamedCol_X" 转换为显示为空。
    """
    if df.empty:
        return ""
    
    # 确保所有列都是字符串类型，以避免替换时出现问题
    # 注意: 此时 DataFrame 应该已经过数值格式化和NA处理，直接转为字符串即可
    df = df.astype(str) 
    
    # 替换标题和数据中的管道符，以避免破坏 Markdown 表格结构
    # 同时将 "UnnamedCol_X" 转换为显示为空
    header_display = []
    for h in df.columns.tolist():
        if isinstance(h, str) and h.startswith("UnnamedCol_"):
            header_display.append("") # 将 UnnamedCol_X 显示为空
        else:
            header_display.append(str(h).replace("|", "\\|"))
    
    rows_values = []
    for _, row_series in df.iterrows():
        rows_values.append([str(cell_val).replace("|", "\\|") for cell_val in row_series.tolist()])
    
    try:
        if table_style == "simple":
            md_table = _rows_to_md_internal(rows_values, header_display)
        elif table_style == "fancy" and hasattr(df, 'to_markdown'):
             # 这需要安装 'tabulate' (pip install tabulate)
            md_table = df.to_markdown(index=False, numalign="left", stralign="left")
        else: # 默认回退到 simple
            md_table = _rows_to_md_internal(rows_values, header_display)
    except Exception as e:
        print(f"警告: 无法使用高级 Markdown 表格格式 ({table_style})。回退到 simple。错误: {e}")
        md_table = _rows_to_md_internal(rows_values, header_display)

    return md_table

def _rows_to_md_internal(rows: list, header: list) -> str:
    """内部辅助函数，将行列表和标题转换为 Markdown 表格字符串。"""
    if not header: 
        return ""
        
    num_cols = len(header)
    # 为标题行生成 Markdown 字符串
    md = "| " + " | ".join([str(h) if pd.notnull(h) else "" for h in header]) + " |\n"
    # 添加分隔线
    md += "| " + " | ".join(['---'] * num_cols) + " |\n"
    
    if not rows: 
        return md

    cleaned_rows_for_md = []
    for row_data_list in rows:
        # 处理空值，并确保行长度与列数匹配
        processed_row = [str(cell) if pd.notnull(cell) else "" for cell in row_data_list]
        if len(processed_row) < num_cols:
            processed_row.extend([""] * (num_cols - len(processed_row)))
        elif len(processed_row) > num_cols:
            processed_row = processed_row[:num_cols]
        cleaned_rows_for_md.append(processed_row)

    # 为数据行生成 Markdown 字符串
    for r_md_row in cleaned_rows_for_md:
        md += "| " + " | ".join(r_md_row) + " |\n"
    return md

def _remove_merged_cell_duplicates_for_display(block_df: pd.DataFrame, block_excel_offset_r: int, block_excel_offset_c: int, merged_cells_full_ranges: list[tuple]) -> pd.DataFrame:
    """
    清理可能包含合并单元格传播值的 DataFrame，
    将非左上角的合并单元格重复值设置为空字符串，以便显示。
    
    Args:
        block_df: 要清理的 DataFrame。
        block_excel_offset_r: 该块在原始工作表中开始的 0-based 行索引。
        block_excel_offset_c: 该块在原始工作表中开始的 0-based 列索引。
        merged_cells_full_ranges: 工作表中所有合并单元格的 (min_r, min_c, max_r, max_c) 元组列表 (0-based)。
    
    Returns:
        清理后的 DataFrame。
    """
    cleaned_df = block_df.copy()
    
    # 遍历 DataFrame 单元格
    for r_df_idx, r_df_series in cleaned_df.iterrows():
        for c_df_idx, cell_value in enumerate(r_df_series):
            # 计算此单元格的原始 Excel 坐标
            # 这些偏移量是相对于整个工作表的块的偏移量
            original_excel_r = block_excel_offset_r + r_df_idx
            original_excel_c = block_excel_offset_c + c_df_idx
            
            is_merged_duplicate = False
            for m_range in merged_cells_full_ranges:
                min_r, min_c, max_r, max_c = m_range
                # 检查单元格是否在合并区域内
                if min_r <= original_excel_r <= max_r and min_c <= original_excel_c <= max_c:
                    # 如果它在合并区域内，并且不是该合并区域的左上角单元格
                    if original_excel_r != min_r or original_excel_c != min_c:
                        is_merged_duplicate = True
                        break 
            
            # 如果是合并单元格的重复项，则设置为空字符串
            if is_merged_duplicate:
                cleaned_df.at[r_df_idx, cleaned_df.columns[c_df_idx]] = "" 
                
    return cleaned_df


def _expand_multiline_cells(df: pd.DataFrame) -> pd.DataFrame:
    """
    将 DataFrame 中包含多行文本的单元格（由换行符 '\n' 分隔）
    展开为 DataFrame 中的多行，以更好地适应 Markdown 表格格式。
    原单元格的第一行内容保留在原行，后续行内容在新增行中。
    其他列在新行中填充空字符串以保持对齐。
    """
    expanded_rows = []
    
    # 遍历 DataFrame 的每一行
    for _, row in df.iterrows():
        # 找出当前行中哪些列包含多行文本
        multiline_cols_data = {col: str(val).split('\n') for col, val in row.items() if '\n' in str(val)}
        
        # 如果当前行没有多行文本单元格，则直接添加原始行
        if not multiline_cols_data:
            expanded_rows.append(row.tolist())
            continue

        # 确定当前行中多行文本单元格的最大行数
        max_lines = max(len(lines) for lines in multiline_cols_data.values())

        # 根据最大行数，为当前逻辑行创建多行
        for i in range(max_lines):
            new_row_data = []
            for col_name in df.columns:
                if col_name in multiline_cols_data:
                    # 获取多行单元格的第 i 行内容，如果超出则为空字符串
                    new_row_data.append(multiline_cols_data[col_name][i] if i < len(multiline_cols_data[col_name]) else "")
                else:
                    # 对于非多行单元格，只有第一行使用其内容，后续行为空字符串
                    new_row_data.append(row[col_name] if i == 0 else "")
            expanded_rows.append(new_row_data)
    
    # 将处理后的行列表转换为新的 DataFrame
    return pd.DataFrame(expanded_rows, columns=df.columns)


def _find_data_blocks(raw_sheet_data: list[list], rows_count: int, cols_count: int, merged_cells_full_ranges: list[tuple]) -> list[tuple]:
    """
    在工作表数据中查找独立的“数据块”。
    使用 BFS 遍历连通组件，并将合并单元格的逻辑连接考虑在内。
    """
    data_blocks = []
    visited_cells = set()

    def has_content(r_idx, c_idx):
        """检查单元格是否有内容，或是否属于有内容的合并单元格。"""
        if not (0 <= r_idx < rows_count and 0 <= c_idx < cols_count):
            return False
        # 确保索引在 raw_sheet_data 的有效范围内
        if r_idx >= len(raw_sheet_data) or c_idx >= len(raw_sheet_data[r_idx]): 
            return False
        
        # 1. 检查单元格本身是否有内容
        if str(raw_sheet_data[r_idx][c_idx]).strip() != "":
            return True

        # 2. 检查它是否是合并单元格的一部分，并且该合并单元格的左上角有内容（使其逻辑上连接）
        for m_range in merged_cells_full_ranges:
            min_r, min_c, max_r, max_c = m_range # 0-based
            if min_r <= r_idx <= max_r and min_c <= c_idx <= max_c:
                # 如果在合并单元格内，并且该合并区域的左上角单元格有内容
                if str(raw_sheet_data[min_r][min_c]).strip() != "":
                    return True
        return False

    for r_start in range(rows_count):
        for c_start in range(cols_count):
            # 如果单元格没有内容或已被访问，则跳过
            if not has_content(r_start, c_start) or (r_start, c_start) in visited_cells:
                continue

            q = [(r_start, c_start)]
            current_block_cells = set([(r_start, c_start)])
            component_min_r, component_max_r = r_start, r_start
            component_min_c, component_max_c = c_start, c_start

            head = 0
            while head < len(q):
                r, c = q[head]
                head += 1
                # 更新当前块的边界
                component_min_r = min(component_min_r, r)
                component_max_r = max(component_max_r, r)
                component_min_c = min(component_min_c, c)
                component_max_c = max(component_max_c, c)

                # 检查所有八个方向的邻居
                for dr_offset in range(-1, 2): 
                    for dc_offset in range(-1, 2): 
                        if dr_offset == 0 and dc_offset == 0: continue # 跳过自身
                        nr, nc = r + dr_offset, c + dc_offset
                        # 如果邻居有内容且未被访问，则添加到队列和当前块中
                        if has_content(nr, nc) and (nr, nc) not in current_block_cells:
                            current_block_cells.add((nr, nc))
                            q.append((nr, nc))
            
            # 将当前块的所有单元格标记为已访问
            visited_cells.update(current_block_cells)
            # 添加当前块的坐标到数据块列表
            data_blocks.append((component_min_r, component_max_r, component_min_c, component_max_c))
    
    # 按行、然后按列对数据块进行排序
    data_blocks.sort(key=lambda b: (b[0], b[2]))
    return data_blocks

def extract_documents_from_excel(excel_path: str, file_name_for_metadata: str) -> list[Document]:
    """
    从 Excel 文件中提取文档。
    遍历每个工作表，识别数据块，并将其转换为 Document 对象。
    """
    all_documents = []
    workbook = None
    is_xls = excel_path.lower().endswith(".xls")
    # 关键字用于识别标题、页脚、总计行等
    group_keywords = ['费用', '欠款', '汇总', '报销', '工资', '社保', '补贴', '补发', '补扣', '补偿', '补缴', '补助', '补款', '总计', '合计', '明细', '列表', '清单', '预算', '决算', '统计', '岗位', '职责', '资格', '说明', '备注', '内容', '要点', '目标', '要求', '经办人', '复核', '审批', '签字'] 
    # 更具体的页脚/总结行关键词，用于判断是否应视为非表格文本块
    footer_keywords = ['合计', '总计', '经办人', '复核', '审批', '签字', '公司', '部门', '制表人', '日期', '签名']
    print(f"开始提取文件: {excel_path}")

    try:
        if is_xls:
            # 如果是 .xls 文件，检查 xlrd 是否可用
            if 'xlrd' not in sys.modules:
                print(f"由于 xlrd 不可用，跳过 .xls 文件 {excel_path}。")
                return []
            try:
                workbook = xlrd.open_workbook(excel_path, formatting_info=True, on_demand=True) 
            except xlrd.XLRDError as e:
                print(f"使用 xlrd 读取 .xls 文件 {excel_path} 时出错: {e}")
                return []
            except Exception as e: 
                print(f"使用 xlrd 打开 .xls 文件 {excel_path} 时发生意外错误: {e}")
                return []
        else: # .xlsx 文件
            # 如果是 .xlsx 文件，检查 openpyxl 是否可用
            if 'openpyxl' not in sys.modules:
                print(f"由于 openpyxl 不可用，跳过 .xlsx 文件 {excel_path}。")
                return []
            try:
                # read_only=False 以便可以访问原始值，data_only=False 确保获取公式而不是结果
                workbook = openpyxl.load_workbook(excel_path, read_only=False, data_only=False) 
            except (InvalidFileException, zipfile.BadZipFile) as e:
                print(f"打开 .xlsx 文件 {excel_path} 时出错: 文件损坏，不是有效的 Excel 文件，或不是 zip 文件。详情: {e}")
                return []
            except Exception as e:
                print(f"打开 .xlsx 文件 {excel_path} 时发生意外错误: {e}")
                return []
        
        if not workbook: return []

        # 获取工作表名称列表
        if is_xls: sheet_names_list = workbook.sheet_names()
        else: sheet_names_list = workbook.sheetnames

        for sheet_idx, sheet_name in enumerate(sheet_names_list):
            print(f"\n--- 正在处理工作表 ({sheet_idx+1}/{len(sheet_names_list)}): {sheet_name} ---")
            raw_sheet_data = []
            cell_meta_info = {} # 存储工作表级别的单元格元数据 (仅针对 .xlsx)
            merged_cells_full_ranges = [] # 存储所有合并单元格的 0-based 范围 (min_r, min_c, max_r, max_c)

            if is_xls:
                try:
                    sheet = workbook.sheet_by_name(sheet_name) 
                except Exception as e:
                    print(f"从 .xls 文件中加载工作表 '{sheet_name}' 时出错: {e}")
                    continue 
                rows_count = sheet.nrows
                cols_count = sheet.ncols if sheet.ncols > 0 else 1 # 确保至少有一列
                # 读取 .xls 文件中的原始数据
                for r in range(rows_count):
                    row_data = []
                    for c in range(cols_count):
                        try:
                            cell = sheet.cell(r, c)
                            cell_value = cell.value
                            # 处理日期类型单元格
                            if cell.ctype == xlrd.XL_CELL_DATE:
                                try:
                                    dt_tuple = xlrd.xldate_as_datetime(cell_value, workbook.datemode)
                                    cell_value = dt_tuple.strftime("%Y-%m-%d %H:%M:%S" if dt_tuple.time() != datetime.time(0,0) else "%Y-%m-%d")
                                except Exception: pass # 日期转换失败则保持原样
                            row_data.append(cell_value)
                        except IndexError: # 处理越界情况
                            row_data.append("") 
                    raw_sheet_data.append(row_data)
                # xlrd 不像 openpyxl 那样容易获取合并单元格信息，所以对于 .xls 文件，merged_cells_full_ranges 将为空。
                # 这意味着 _find_data_blocks 将仅依赖 raw_sheet_data 的内容，这可能仍会导致稀疏表格被分割。
            else: # .xlsx 文件
                sheet = workbook[sheet_name]
                rows_count = sheet.max_row if sheet.max_row else 0
                cols_count = sheet.max_column if sheet.max_column else 0
                if rows_count == 0 or cols_count == 0: 
                    print(f"工作表 '{sheet_name}' 在 .xlsx 文件中似乎为空。")
                    continue 
                # 初始化 raw_sheet_data 为空字符串，以便按索引填充
                raw_sheet_data = [['' for _ in range(cols_count)] for _ in range(rows_count)]
                # 遍历 .xlsx 文件中的单元格并填充 raw_sheet_data 和 cell_meta_info
                for r_idx, row_iter in enumerate(sheet.iter_rows(min_row=1, max_row=rows_count, min_col=1, max_col=cols_count)):
                    for c_idx, cell in enumerate(row_iter):
                        if not isinstance(cell, OpenpyxlCell): # 确保是有效的单元格对象
                             raw_sheet_data[r_idx][c_idx] = "" 
                             continue
                        raw_sheet_data[r_idx][c_idx] = cell.value # 这处理了合并单元格，只有左上角有值
                        
                        meta_entry = {}
                        # 提取公式
                        if hasattr(cell, 'data_type') and cell.data_type == 'f': 
                             if isinstance(cell.value, str) and cell.value.startswith('='):
                                 meta_entry['formula'] = cell.value
                        # 提取注释
                        if cell.comment: meta_entry['comment'] = str(cell.comment.text)
                        # 提取格式信息（粗体、斜体）
                        if cell.font: 
                            formats = []
                            if cell.font.bold: formats.append('Bold')
                            if cell.font.italic: formats.append('Italic')
                            if formats: meta_entry['format'] = ", ".join(formats)
                        # 如果存在元数据，则存储到工作表级别的 cell_meta_info
                        if meta_entry:
                            cell_meta_info[(r_idx, c_idx)] = meta_entry 
                
                # 为 openpyxl 收集合并单元格范围
                for merged_range in sheet.merged_cells.ranges:
                    min_col, min_r, max_col, max_r = merged_range.bounds 
                    # 存储 0-based 坐标
                    merged_cells_full_ranges.append((min_r -1, min_col -1, max_r -1, max_col -1)) 

            if not raw_sheet_data:
                print(f"读取工作表 '{sheet_name}' 后没有数据。")
                continue

            # 使用 _find_data_blocks 识别数据块
            data_blocks_coords = _find_data_blocks(raw_sheet_data, rows_count, cols_count, merged_cells_full_ranges)
            print(f"在工作表 '{sheet_name}' 中识别到 {len(data_blocks_coords)} 个潜在数据块。")

            # 新增: 后处理，用于合并垂直相邻/重叠的、看起来属于同一表格的数据块
            merged_data_blocks = []
            if data_blocks_coords:
                current_block_merged = list(data_blocks_coords[0]) # (min_r, max_r, min_c, max_c)
                
                for i in range(1, len(data_blocks_coords)):
                    next_block = list(data_blocks_coords[i])
                    
                    # 检查垂直接近度和列重叠
                    consider_merge = False
                    
                    vertical_distance = next_block[0] - current_block_merged[1] - 1 # 两个块之间的空行数
                    
                    # 计算列重叠
                    overlap_min_c = max(current_block_merged[2], next_block[2])
                    overlap_max_c = min(current_block_merged[3], next_block[3])
                    
                    overlap_width = overlap_max_c - overlap_min_c + 1
                    current_block_width = current_block_merged[3] - current_block_merged[2] + 1
                    next_block_width = next_block[3] - next_block[2] + 1
                    
                    # 条件1: 垂直距离合理 (允许最多5个空行，特别是对于长的、连续的账目类表格)
                    if vertical_distance >= 0 and vertical_distance <= 5: 
                        # 条件2: 列对齐或显著重叠
                        # 如果列起始/结束位置非常接近 (例如，相差不超过1列)
                        # 或者有非常高的列重叠度 (例如，超过75%的较窄块的列宽)
                        # 或者一个块完全包含另一个块 (常见于主表和其页脚/总计行)
                        col_start_diff = abs(current_block_merged[2] - next_block[2])
                        col_end_diff = abs(current_block_merged[3] - next_block[3])

                        if (col_start_diff <= 1 and col_end_diff <= 1) or \
                           (overlap_width / max(1, current_block_width, next_block_width) > 0.75) or \
                           (current_block_merged[2] <= next_block[2] and current_block_merged[3] >= next_block[3]) or \
                           (next_block[2] <= current_block_merged[2] and next_block[3] >= current_block_merged[3]):
                            consider_merge = True

                    if consider_merge:
                        # 合并块
                        current_block_merged[1] = max(current_block_merged[1], next_block[1]) # 更新最大行
                        current_block_merged[2] = min(current_block_merged[2], next_block[2]) # 更新最小列
                        current_block_merged[3] = max(current_block_merged[3], next_block[3]) # 更新最大列
                        current_block_merged[0] = min(current_block_merged[0], next_block[0]) # 更新最小行 (确保块起始行是最早的)
                    else: # 没有显著重叠或不够接近，所以这是一个新的独立块
                        merged_data_blocks.append(tuple(current_block_merged))
                        current_block_merged = list(next_block)
                
                merged_data_blocks.append(tuple(current_block_merged)) # 添加最后一个块
                data_blocks_coords = merged_data_blocks
                print(f"合并后，工作表 '{sheet_name}' 中有 {len(data_blocks_coords)} 个处理过的数据块。")

            # 合并后再次排序，以防万一 (通常会保持顺序)
            data_blocks_coords.sort(key=lambda b: (b[0], b[2]))

            sheet_documents = []
            # 遍历每个数据块并进行处理
            for block_idx, (min_r_block, max_r_block, min_c_block, max_c_block) in enumerate(data_blocks_coords):
                # 初始化当前数据块的单元格元数据，解决 NameError
                block_cell_metadata = {}
                if not is_xls: # 仅对 .xlsx 文件收集的元数据进行筛选
                    for (r_orig, c_orig), meta_val in cell_meta_info.items():
                        # 检查原始单元格坐标是否落在当前数据块的范围内
                        if min_r_block <= r_orig <= max_r_block and min_c_block <= c_orig <= max_c_block:
                            # 存储原始坐标，因为更直接
                            block_cell_metadata[(r_orig, c_orig)] = meta_val

                # 提取原始的块数据
                block_data_raw = [raw_sheet_data[r_idx][min_c_block : max_c_block + 1] for r_idx in range(min_r_block, max_r_block + 1)]
                if not block_data_raw: continue
                block_df = pd.DataFrame(block_data_raw)
                if block_df.empty: continue

                # 步骤1: 清理合并单元格的重复项以便显示
                block_df = _remove_merged_cell_duplicates_for_display(block_df, min_r_block, min_c_block, merged_cells_full_ranges)
                
                # 步骤2: 展开包含多行文本的单元格
                block_df = _expand_multiline_cells(block_df)

                # 将所有空字符串替换为 Pandas 的 NA 值，以便 dropna 能够识别
                block_df = block_df.replace('', pd.NA) 
                # 删除所有值都为 NA 的行和列
                block_df.dropna(how='all', axis=0, inplace=True) 
                block_df.dropna(how='all', axis=1, inplace=True) 
                block_df.reset_index(drop=True, inplace=True)
                if block_df.empty: continue
                # 此时再将 NA 填充回空字符串，因为 Markdown 表格需要字符串
                block_df = block_df.fillna("") 

                # --- 启发式判断该块是表格还是纯文本 ---
                is_likely_table = True
                header_row_idx = -1 # 初始化为 -1，表示未找到表头

                # 尝试在 block_df 的前几行中找到表头 (最多检查前 5 行)
                # 表头通常具有以下特征：
                # 1. 较多的非空单元格 (例如，超过列数的一半，或者至少 3 个非空单元格)
                # 2. 组合起来的文本不应过长 (不像一段描述性文字)
                for r_idx_potential_header in range(min(5, len(block_df))): 
                    potential_header_row_vals = block_df.iloc[r_idx_potential_header]
                    num_non_empty_header_cells = sum(1 for x in potential_header_row_vals if str(x).strip())
                    
                    if (num_non_empty_header_cells >= len(potential_header_row_vals) / 2) or \
                       (num_non_empty_header_cells >= 3 and len(potential_header_row_vals) > 0) :
                        combined_header_str = " ".join([str(x).strip() for x in potential_header_row_vals if str(x).strip()])
                        # 启发式: 假设表头不会太长 (例如，小于 100 个字符)
                        if len(combined_header_str) < 100: 
                            header_row_idx = r_idx_potential_header
                            break # 找到表头，跳出循环
                
                if header_row_idx == -1: # 如果没有找到明确的表头
                    is_likely_table = False 
                    # 附加检查: 如果列数很多且行数也很多，且非空单元格密度较高，即使没有明确表头，也可能是一个表格。
                    # 这里放宽列数限制，因为有些表格可能只有少量关键列但仍是表格
                    if len(block_df.columns) >= 2 and len(block_df) > 3: # 至少2列且行数较多
                        non_empty_cells_count = block_df.map(lambda x: bool(str(x).strip())).sum().sum() # 修复 applymap warning
                        total_cells_in_df = block_df.shape[0] * block_df.shape[1]
                        if total_cells_in_df > 0 and (non_empty_cells_count / total_cells_in_df > 0.4): # 如果超过 40% 的单元格有内容
                            is_likely_table = True # 认为是表格，即使没有显式表头
                            # (如果 is_likely_table 重新变为 True，则 header_row_idx 仍为 -1，后面会分配 Col_X 名称)

                # 覆盖: 如果块中包含常见的页脚关键词 (如 '合计', '经办人') 并且它是一个窄块（列数少于5）
                # 那么它可能是一个文本块或总结，而不是主要表格。
                # 备注：此处仅作为辅助判断，最终的"合计"行不再被特殊移除
                # 使用 DataFrame.map 替代 applymap
                is_footer_like_block = block_df.map(lambda x: any(kw in str(x).lower() for kw in footer_keywords)).any().any()
                if is_footer_like_block and len(block_df.columns) < 5:
                    is_likely_table = False 

                # --- 如果不是表格，则按文本块处理 (例如，标题、描述、页脚) ---
                if not is_likely_table:
                    text_content_parts = []
                    for r_idx in range(len(block_df)):
                        row_cells = [str(x).strip() for x in block_df.iloc[r_idx]]
                        # 仅当行中有实际内容时才包含
                        if any(c for c in row_cells): # 检查行中是否有任何单元格有内容
                            # 将行中的非空单元格用空格连接，并去除首尾空白
                            text_content_parts.append(" ".join([c for c in row_cells if c]).strip())
                    
                    final_text_content = "\n\n".join([p for p in text_content_parts if p]).strip() 

                    if final_text_content.strip(): # 只有当最终文本内容不为空时才生成文档
                        text_block_title = f"文本块 {block_idx + 1}"
                        # 尝试从块内容中推断标题
                        first_valid_row_text = ""
                        for r_idx in range(len(block_df)):
                            row_vals = [str(x).strip() for x in block_df.iloc[r_idx] if str(x).strip()]
                            if row_vals:
                                first_valid_row_text = " ".join(row_vals)
                                break
                        
                        if first_valid_row_text:
                            if len(first_valid_row_text) < 50: # 如果足够短，用作标题
                                text_block_title = first_valid_row_text
                            else: # 否则，尝试取前几个词
                                first_words = " ".join(first_valid_row_text.split()[:5])
                                if len(first_words) < 50 and first_words:
                                    text_block_title = first_words
                        
                        document_metadata = {
                            "file_name": file_name_for_metadata, "sheet_name": sheet_name,
                            "block_title": text_block_title, "block_index_in_sheet": block_idx + 1,
                            "excel_range": f"R{min_r_block+1}C{min_c_block+1}:R{max_r_block+1}C{max_c_block+1}",
                            "source_type": "excel_structured_text_block"
                        }
                        # 使用 H3 作为文本块的标题，以便更好的层级结构
                        sheet_documents.append(Document(page_content=f"### {text_block_title}\n{final_text_content}", metadata=document_metadata))
                    continue # 跳过当前块，继续处理下一个数据块

                # --- 如果识别为表格，则继续表格处理 ---
                if header_row_idx != -1: # 如果找到了表头行
                    headers_list = [str(h).strip() for h in block_df.iloc[header_row_idx]]
                    seen_headers_count = {}; final_headers_list = []
                    # 处理重复的表头名称
                    for h_val_idx, h_val_content in enumerate(headers_list):
                        _h_val_content = h_val_content if h_val_content else f"UnnamedCol_{h_val_idx}"
                        if _h_val_content in seen_headers_count:
                            seen_headers_count[_h_val_content] += 1
                            final_headers_list.append(f"{_h_val_content}_{seen_headers_count[_h_val_content]}")
                        else: 
                            seen_headers_count[_h_val_content] = 0 # 初始化为 0，以便第一次出现时是原始名称
                            final_headers_list.append(_h_val_content)
                    block_df.columns = final_headers_list
                    block_df = block_df.iloc[header_row_idx+1:].reset_index(drop=True) # 删除表头行
                else: # 没有明确的表头，使用 Col_X 名称
                    block_df.columns = [f"Col_{j}" for j in range(len(block_df.columns))]

                # 移除处理备注列的合并和摘要行删除的逻辑
                # 原来的逻辑导致 "合计" 等行被不当合并或删除，现在将其作为普通行保留在表格中。
                
                # --- 新增: 统一数值格式 ---
                # 在转换为 Markdown 之前，格式化数值列
                for col in block_df.columns:
                    # 将空字符串替换为 NaN，以便 pd.to_numeric 处理
                    temp_series = pd.to_numeric(block_df[col].replace('', pd.NA), errors='coerce')
                    
                    # 检查是否是数值类型，且该列包含非空值
                    if not temp_series.isna().all():
                        # 判断是否包含浮点数或者列名暗示为金额 (宽松判断，以便格式化如 "余额", "借方", "贷方")
                        contains_float_values = temp_series.apply(lambda x: isinstance(x, float) and x % 1 != 0 and pd.notna(x)).any()
                        is_currency_like_col_name = any(kw in str(col).lower() for kw in ['余额', '借方', '贷方', '金额', '总计', '合计'])
                        
                        if contains_float_values or is_currency_like_col_name:
                            # 舍入到两位小数，并添加千位分隔符
                            block_df[col] = temp_series.round(2).apply(
                                lambda x: f"{x:,.2f}" if pd.notna(x) else ""
                            )
                        else:
                            # 对于整数或其他非浮点性质的数字，不加小数，只加千位分隔符
                            block_df[col] = temp_series.apply(
                                lambda x: f"{int(x):,}" if pd.notna(x) else ""
                            )
                
                # 确保所有单元格最终都是字符串，以便 Markdown 渲染
                block_df = block_df.astype(str)

                # 生成数据块的标题和最终页面内容
                effective_block_title = ""
                final_page_content_md = ""

                # 尝试从当前块上方的一行中推断标题 (通用逻辑)
                inferred_title_from_above = None
                if min_r_block > 0: 
                    title_candidate_src_row = raw_sheet_data[min_r_block -1]
                    potential_block_title_parts = []
                    last_non_empty_col = -1
                    # 尝试将相邻的、非空字符串合并为标题
                    for c_content_idx in range(min_c_block, min(max_c_block + 1, len(title_candidate_src_row))):
                        cell_val = str(title_candidate_src_row[c_content_idx]).strip()
                        if cell_val:
                            # 如果是连续的非空单元格，则拼接
                            if potential_block_title_parts and c_content_idx == last_non_empty_col + 1:
                                potential_block_title_parts[-1] += " " + cell_val
                            else: # 否则，开始一个新的部分
                                potential_block_title_parts.append(cell_val)
                            last_non_empty_col = c_content_idx
                    
                    if potential_block_title_parts:
                        _inferred = " ".join(potential_block_title_parts).strip()
                        # 检查推断标题是否合理，例如长度适中，包含关键词，或有冒号等（如“岗位职责：”）
                        if _inferred and len(_inferred) < 100 and \
                           (any(kw in _inferred for kw in group_keywords) or 
                            (len(potential_block_title_parts) == 1 and (_inferred.endswith(('：')) or _inferred.endswith(':')))): 
                            inferred_title_from_above = _inferred

                md_table_content = _df_to_md(block_df)

                # 根据条件决定是否添加块标题和其级别
                if inferred_title_from_above:
                    effective_block_title = inferred_title_from_above
                    final_page_content_md = f"#### {effective_block_title}\n{md_table_content}"
                elif block_idx == 0 and not inferred_title_from_above: # 如果是该工作表第一个数据块，且没有推断出明确标题
                    # 此时逻辑标题为工作表名，不在内容中重复，直接输出表格
                    effective_block_title = sheet_name 
                    final_page_content_md = md_table_content 
                else: # 如果是后续数据块，且没有推断出明确标题
                    effective_block_title = f"数据块 {block_idx + 1}"
                    final_page_content_md = f"#### {effective_block_title}\n{md_table_content}"
                
                # 构建 Document 对象的元数据
                document_metadata = {
                    "file_name": file_name_for_metadata, "sheet_name": sheet_name,
                    "block_title": effective_block_title, # 使用最终确定的标题
                    "block_index_in_sheet": block_idx + 1,
                    "excel_range": f"R{min_r_block+1}C{min_c_block+1}:R{max_r_block+1}C{max_c_block+1}",
                    "headers": block_df.columns.tolist(), # 将识别到的表头添加到元数据中
                    "source_type": "excel_structured_cli"
                }
                # 如果当前数据块有单元格元数据，则添加
                if block_cell_metadata: # 这里的 block_cell_metadata 现在肯定被初始化了
                    document_metadata["cell_metadata"] = block_cell_metadata
                
                sheet_documents.append(Document(page_content=final_page_content_md, metadata=document_metadata))
            all_documents.extend(sheet_documents)
            print(f"工作表 '{sheet_name}' 处理完成，找到 {len(sheet_documents)} 个文档部分。")
        
        # 对于 xlrd，释放资源
        if is_xls and workbook and hasattr(workbook, 'release_resources'):
            workbook.release_resources() 

    except Exception as e:
        print(f"处理文件 {excel_path} 时发生严重错误: {e}")
        import traceback
        traceback.print_exc() # 打印详细的错误堆栈信息
    
    print(f"文件 {excel_path} 提取完成。总文档数: {len(all_documents)}")
    return all_documents

def process_single_file(excel_file_path: str):
    """
    处理单个 Excel 文件，将其解析为 Markdown 文件。
    """
    print(f"\n正在处理单个文件: {excel_file_path}")
    file_base_name = os.path.splitext(os.path.basename(excel_file_path))[0]
    output_md_path = os.path.join(os.path.dirname(excel_file_path), file_base_name + ".md")
    
    # 提取文档
    documents_list = extract_documents_from_excel(excel_file_path, file_base_name)
    if not documents_list:
        print(f"未从 {excel_file_path} 提取到任何内容。未创建 Markdown 文件。")
        return
    
    # 按工作表分组文档内容
    md_content_by_sheet = {}
    encountered_sheet_order = [] # 保持工作表的原始顺序
    for doc_item in documents_list:
        sheet_name_meta = doc_item.metadata.get("sheet_name", "未知工作表")
        if sheet_name_meta not in md_content_by_sheet:
            md_content_by_sheet[sheet_name_meta] = []
            encountered_sheet_order.append(sheet_name_meta)
        md_content_by_sheet[sheet_name_meta].append(doc_item.page_content) 
    
    # 拼接最终的 Markdown 内容
    final_md_parts_list = []
    for sheet_name_ordered in encountered_sheet_order: 
        # 使用 H1 作为工作表标题，并去除前缀，模仿 output (1).md
        final_md_parts_list.append(f"# {sheet_name_ordered}\n") 
        final_md_parts_list.extend(md_content_by_sheet[sheet_name_ordered])
        # 在每个工作表内容末尾添加水平分隔线
        final_md_parts_list.append("\n---\n") # 每个工作表末尾添加一个水平分隔符
    
    final_md_output_str = "\n".join(final_md_parts_list).strip()

    # 写入 Markdown 文件
    if final_md_output_str:
        try:
            with open(output_md_path, "w", encoding="utf-8") as f: 
                f.write(final_md_output_str)
            print(f"成功将 Markdown 写入: {output_md_path}")
        except Exception as e: 
            print(f"将 Markdown 写入 {output_md_path} 时出错: {e}")
    else: 
        print(f"没有要为 {excel_file_path} 写入的重要内容，未创建 Markdown 文件。")

def process_directory(excel_dir_path: str):
    """
    处理指定目录下的所有 Excel 文件，将其解析为 Markdown 文件，并保存到新的目录结构中。
    """
    print(f"\n正在处理目录: {excel_dir_path}")
    # 获取父目录和当前目录的基础名称
    parent_dir_path = os.path.dirname(os.path.abspath(excel_dir_path.rstrip("/\\")))
    dir_base_name = os.path.basename(os.path.abspath(excel_dir_path.rstrip("/\\")))
    # 生成带时间戳的输出目录名称
    timestamp_now = datetime.datetime.now().strftime("%Y%m%d-%H%M")
    md_files_root_dir = os.path.join(parent_dir_path, f"{dir_base_name}-md解析-{timestamp_now}")
    
    # 创建输出根目录
    if not os.path.exists(md_files_root_dir): 
        os.makedirs(md_files_root_dir)
    print(f"Markdown 输出将保存到: {md_files_root_dir}")

    # 遍历目录中的所有文件
    for current_root, _, files_in_root in os.walk(excel_dir_path):
        for file_item_name in files_in_root:
            # 跳过临时文件
            if file_item_name.startswith("~$") or file_item_name.startswith(".~"):
                print(f"跳过临时文件: {os.path.join(current_root, file_item_name)}")
                continue
            # 检查文件是否是 Excel 文件
            if file_item_name.lower().endswith((".xlsx", ".xls")): 
                excel_file_full_path = os.path.join(current_root, file_item_name)
                # 计算相对于输入目录的相对路径
                relative_path_to_input_dir = os.path.relpath(excel_file_full_path, excel_dir_path)
                # 构建输出 Markdown 文件的路径
                output_md_file_name = os.path.splitext(os.path.basename(relative_path_to_input_dir))[0] + ".md"
                output_md_file_target_dir = os.path.join(md_files_root_dir, os.path.dirname(relative_path_to_input_dir))
                
                # 创建输出子目录
                if not os.path.exists(output_md_file_target_dir): 
                    os.makedirs(output_md_file_target_dir)
                output_md_full_path = os.path.join(output_md_file_target_dir, output_md_file_name)

                print(f"\n正在处理: {excel_file_full_path}")
                file_base_name_for_metadata = os.path.splitext(file_item_name)[0]
                
                # 提取文档
                documents_list_from_file = extract_documents_from_excel(excel_file_full_path, file_base_name_for_metadata)
                if not documents_list_from_file:
                    print(f"未从 {excel_file_full_path} 提取到任何内容。未创建 Markdown 文件。")
                    continue
                
                # 按工作表分组文档内容
                md_content_by_sheet_for_dir = {}
                encountered_sheet_order_for_dir = []
                for doc_item_dir in documents_list_from_file:
                    sheet_name_meta_dir = doc_item_dir.metadata.get("sheet_name", "未知工作表")
                    if sheet_name_meta_dir not in md_content_by_sheet_for_dir:
                        md_content_by_sheet_for_dir[sheet_name_meta_dir] = []
                        encountered_sheet_order_for_dir.append(sheet_name_meta_dir)
                    md_content_by_sheet_for_dir[sheet_name_meta_dir].append(doc_item_dir.page_content)
                
                # 拼接最终的 Markdown 内容
                final_md_parts_list_for_dir = []
                for sheet_name_ordered_dir in encountered_sheet_order_for_dir:
                    final_md_parts_list_for_dir.append(f"# {sheet_name_ordered_dir}\n")
                    final_md_parts_list_for_dir.extend(md_content_by_sheet_for_dir[sheet_name_ordered_dir])
                    final_md_parts_list_for_dir.append("\n---\n") # 每个工作表末尾添加一个水平分隔符
                final_md_output_str_for_dir = "\n".join(final_md_parts_list_for_dir).strip()
                
                # 写入 Markdown 文件
                if final_md_output_str_for_dir:
                    try:
                        with open(output_md_full_path, "w", encoding="utf-8") as f: 
                            f.write(final_md_output_str_for_dir)
                        print(f"成功将 Markdown 写入: {output_md_full_path}")
                    except Exception as e: 
                        print(f"将 Markdown 写入 {output_md_full_path} 时出错: {e}")
                else: 
                    print(f"没有要为 {excel_file_full_path} 写入的重要内容，未创建 Markdown 文件。")
    print(f"\n目录中所有文件处理完成。输出在: {md_files_root_dir}")

def main():
    """主函数，处理命令行参数，调用文件或目录处理函数。"""
    if len(sys.argv) != 2:
        print("用法: python i_excel_md_convert.py <excel_文件路径_或_目录路径>")
        sys.exit(1)
    input_path = sys.argv[1]
    if not os.path.exists(input_path):
        print(f"错误: 输入路径不存在: {input_path}")
        sys.exit(1)
    
    if os.path.isfile(input_path):
        # 检查是否是有效的 Excel 文件且不是临时文件
        if (input_path.lower().endswith((".xlsx", ".xls"))) and \
           not os.path.basename(input_path).startswith("~$") and \
           not os.path.basename(input_path).startswith(".~"):
            process_single_file(input_path)
        else: 
            print("错误: 输入文件不是有效的 .xlsx 或 .xls 文件 (或它可能是临时文件)。")
            sys.exit(1)
    elif os.path.isdir(input_path): 
        process_directory(input_path)
    else: 
        print("错误: 输入路径不是有效的文件或目录。")
        sys.exit(1)

if __name__ == "__main__":
    main()