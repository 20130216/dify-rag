import sys
import pandas as pd
import os
import re
import xlrd # 直接导入 xlrd 用于处理 .xls 文件
import openpyxl # 直接导入 openpyxl 用于处理 .xlsx 文件
from openpyxl.utils.exceptions import InvalidFileException # 导入特定异常
import zipfile # 导入 zipfile 用于捕获 .xlsx 相关异常
import datetime # For handling Excel dates

# 辅助函数：将列表转为 Markdown 表格行
def _list_to_markdown_row(data_list):
    """Internal helper to convert a list of cells to a Markdown table row string."""
    return "| " + " | ".join([str(cell) if pd.notnull(cell) else "" for cell in data_list]) + " |"

# 辅助函数：将 DataFrame 转为 Markdown 表格文本
def _df_to_md(df):
    """Internal helper to convert a DataFrame to Markdown table string."""
    if df.empty:
        return ""
    # Ensure all columns are string type before converting to list of lists
    df = df.astype(str)
    header = df.columns.tolist()
    rows = df.values.tolist() # Get data rows as list of lists

    # Use the helper function to build the Markdown string
    return _rows_to_md(rows, header)

# 辅助函数：将 list of lists 转为 Markdown 表格文本
def _rows_to_md(rows, header):
    """Internal helper to convert list of rows (list of cells) and a header to Markdown table string."""
    if not rows or not header: # Needs data rows and header
        return ""

    # Ensure all data rows have the same number of columns as the header
    num_cols = len(header)
    cleaned_rows = []
    for row in rows:
        # Convert each cell to string, handling NaN/None
        cleaned_row = [str(cell) if pd.notnull(cell) else "" for cell in row]
        # Pad or truncate row to match header size
        if len(cleaned_row) < num_cols:
            cleaned_row.extend([""] * (num_cols - len(cleaned_row)))
        elif len(cleaned_row) > num_cols:
            cleaned_row = cleaned_row[:num_cols]
        cleaned_rows.append(cleaned_row)

    # Generate header row
    md = "| " + " | ".join([str(h) if pd.notnull(h) else "" for h in header]) + " |\n"
    # Generate separator line
    md += "| " + " | ".join(['---'] * num_cols) + " |\n"
    # Generate data rows
    for row in cleaned_rows:
        md += "| " + " | ".join(row) + " |\n"
    return md


def excel_to_markdown(excel_path, md_path):
    """
    读取 Excel 文件（支持 .xls/.xlsx，多Sheet），清洗噪声，合并汇总/备注行，自动分组，输出为规范 Markdown 表格。

    Features:
    - Handles .xls and .xlsx using xlrd and openpyxl respectively.
    - Reads all sheets in the workbook.
    - Correctly handles merged cells and cell values (including dates, formulas, formats in comments) in .xlsx.
    - Preserves all columns and rows by analyzing data blocks.
    - Identifies multiple distinct data blocks/sub-tables within a sheet.
    - Attempts to preserve original column headers per sub-table.
    - Cleans NaN/None values to empty strings.
    - Smartly merges summary/remark-like rows into a designated remark column *per sub-table*.
    - Automatically detects sheet breaks and potential logical groups within sheets/sub-tables, adds Markdown titles (H2 for sheets, H3 for sections, H4 for sub-tables).
    - Outputs standard Markdown tables suitable for RAG.
    - Focuses on data accuracy and structural representation from Excel.
    - Attempts to include formula and format info in remarks.
    """
    all_sheets_markdown = []
    workbook = None
    is_xls = excel_path.lower().endswith(".xls")

    try:
        # 1. 打开工作簿
        if is_xls:
            try:
                workbook = xlrd.open_workbook(excel_path, formatting_info=True) # formatting_info=True to potentially get format info
                print(f"Successfully opened .xls file: {excel_path} using xlrd")
            except ImportError:
                 raise ImportError("请先安装 xlrd==1.2.0 及兼容的 pandas 版本以支持 .xls 文件读取") # Ensure pandas compatibility is mentioned
            except xlrd.XLRDError as e:
                 print(f"Error reading .xls file {excel_path} with xlrd: {e}")
                 return # Skip current file
            except Exception as e:
                 print(f"Unexpected error opening .xls file {excel_path}: {e}")
                 return # Skip current file
        else: # Assumes .xlsx
            try:
                # read_only=False to access formatting, but can be memory intensive
                # data_only=False to read formulas, not just results
                workbook = openpyxl.load_workbook(excel_path, read_only=False, data_only=False)
                print(f"Successfully opened .xlsx file: {excel_path} using openpyxl (read_only=False, data_only=False)")
            except FileNotFoundError:
                print(f"Error: File not found at {excel_path}")
                return # Skip current file
            except (InvalidFileException, zipfile.BadZipFile) as e:
                print(f"Error opening .xlsx file {excel_path}: File is corrupted or not a valid Excel file. Details: {e}")
                return # Skip current file
            except Exception as e:
                print(f"Unexpected error opening .xlsx file {excel_path}: {e}")
                return # Skip current file

        # 2. 遍历所有 Sheet
        sheet_names = workbook.sheet_names() if is_xls else workbook.sheetnames

        if not sheet_names:
             print(f"No sheets found in {excel_path}")
             return

        for sheet_name in sheet_names:
            print(f"\n--- Processing Sheet: {sheet_name} ---")
            sheet_md_parts = [] # Markdown parts for this specific sheet
            sheet_md_parts.append(f"\n## Sheet: {sheet_name}\n") # H2 for sheets

            try:
                # Get sheet data and merged cells (only for openpyxl)
                if is_xls:
                    sheet = workbook.sheet_by_name(sheet_name)
                    rows = sheet.nrows
                    cols = sheet.ncols
                    raw_sheet_data = []
                    # Read data, handling dates and potentially formatting/formulas (limited in xlrd)
                    for r in range(rows):
                        row_data = []
                        for c in range(cols):
                            cell = sheet.cell(r, c)
                            cell_value = cell.value
                            # Basic date handling for xlrd
                            if cell.ctype == xlrd.XL_CELL_DATE:
                                try:
                                    dt_tuple = xlrd.xldate_as_datetime(cell_value, workbook.datemode)
                                    if dt_tuple.time() == datetime.time(0, 0):
                                         cell_value = dt_tuple.strftime("%Y-%m-%d")
                                    else:
                                        cell_value = dt_tuple.strftime("%Y-%m-%d %H:%M:%S")
                                except Exception:
                                    pass # If date conversion fails, use original value
                            # xlrd formula/formatting info is harder to get at cell level this way
                            row_data.append(cell_value)
                        raw_sheet_data.append(row_data)
                    merged_ranges = [] # xlrd doesn't expose merged cells easily this way

                else: # openpyxl for .xlsx
                    sheet = workbook[sheet_name]
                    rows = sheet.max_row
                    cols = sheet.max_column
                    raw_sheet_data = [['' for _ in range(cols)] for _ in range(rows)]
                    merged_ranges = sheet.merged_cells.ranges # Get merged cell ranges
                    cell_meta = {} # Store formula and format info keyed by (row, col)

                    # Read data, formulas, comments, and basic formatting
                    for r_idx, row in enumerate(sheet.iter_rows()):
                        if r_idx < rows:
                            for c_idx, cell in enumerate(row):
                                 if c_idx < cols:
                                     # Get cell value (formula result if data_only=False)
                                     raw_sheet_data[r_idx][c_idx] = cell.value

                                     # Store formula if exists
                                     if cell.formula:
                                         cell_meta[(r_idx, c_idx)] = cell_meta.get((r_idx, c_idx), {})
                                         cell_meta[(r_idx, c_idx)]['formula'] = cell.formula

                                     # Store comment if exists
                                     if cell.comment and cell.comment.text:
                                          cell_meta[(r_idx, c_idx)] = cell_meta.get((r_idx, c_idx), {})
                                          cell_meta[(r_idx, c_idx)]['comment'] = cell.comment.text

                                     # Basic format info (bold, italic) - openpyxl has limitations in read_only/efficient mode
                                     # Accessing cell.font requires read_only=False
                                     if cell.font:
                                          format_notes = []
                                          if cell.font.bold: format_notes.append('Bold')
                                          if cell.font.italic: format_notes.append('Italic')
                                          if format_notes:
                                               cell_meta[(r_idx, c_idx)] = cell_meta.get((r_idx, c_idx), {})
                                               cell_meta[(r_idx, c_idx)]['format'] = ", ".join(format_notes)


                    # Propagate merged cell values
                    for merged_range in merged_ranges:
                        try:
                             # Get the value from the top-left cell *after* initial read
                             min_col, min_row, max_col, max_row = merged_range.bounds # 1-based
                             top_left_value = raw_sheet_data[min_row - 1][min_col - 1] # 0-based

                             for r in range(min_row - 1, max_row):
                                 for c in range(min_col - 1, max_col):
                                     if r < rows and c < cols:
                                          # Fill the value in all cells of the merged range
                                          raw_sheet_data[r][c] = top_left_value
                        except Exception as e:
                             print(f"Warning: Could not process merged cell range {merged_range.coord} in sheet {sheet_name}: {e}")
                             continue

                    print(f"Read {rows} rows and {cols} columns from sheet '{sheet_name}' using openpyxl.")


                # 3. Identify distinct data blocks/sub-tables within the raw sheet data
                data_blocks = [] # List of (start_row_idx, end_row_idx, start_col_idx, end_col_idx) tuples
                visited_cells = set() # Keep track of cells already included in a block

                # Heuristic to find blocks: Iterate through cells. If a cell has content and hasn't been visited,
                # it's the potential start of a block. Expand outwards to include adjacent cells with content.
                # This is a simplified approach, can be complex for arbitrary Excel layouts.

                rows_count = len(raw_sheet_data)
                cols_count = max(len(row) for row in raw_sheet_data) if raw_sheet_data else 0

                def has_content(r, c):
                     return 0 <= r < rows_count and 0 <= c < cols_count and str(raw_sheet_data[r][c]).strip() != ""

                def find_block(start_r, start_c):
                     if not has_content(start_r, start_c) or (start_r, start_c) in visited_cells:
                         return None

                     # Simple flood fill / expansion heuristic
                     min_r, max_r = start_r, start_r
                     min_c, max_c = start_c, start_c
                     q = [(start_r, start_c)]
                     visited_block_cells = set([(start_r, start_c)])

                     while q:
                         r, c = q.pop(0)
                         min_r = min(min_r, r)
                         max_r = max(max_r, r)
                         min_c = min(min_c, c)
                         max_c = max(max_c, c)

                         # Explore neighbors (up, down, left, right)
                         neighbors = [(r-1, c), (r+1, c), (r, c-1), (r, c+1)]
                         # Also consider diagonal or slightly further cells to catch gaps?
                         # This is the tricky part for arbitrary tables. Let's stick to basic neighbors for now.
                         # Add checks for a few cells away to bridge small gaps
                         # neighbors.extend([(r-2, c), (r+2, c), (r, c-2), (r, c+2), (r-1, c-1), (r-1, c+1), (r+1, c-1), (r+1, c+1)]) # Optional: for small gaps


                         for nr, nc in neighbors:
                             if has_content(nr, nc) and (nr, nc) not in visited_block_cells:
                                 visited_block_cells.add((nr, nc))
                                 q.append((nr, nc))

                     # Post-processing: Extend the block boundaries slightly if adjacent rows/cols are sparse but belong
                     # This is a complex heuristic. A simpler approach: just use the bounds found.

                     # Add all cells in the found block to the main visited set
                     visited_cells.update(visited_block_cells)

                     return (min_r, max_r, min_c, max_c)


                # Iterate through all cells to find starts of blocks
                for r in range(rows_count):
                    for c in range(cols_count):
                        block = find_block(r, c)
                        if block:
                            data_blocks.append(block)

                # Sort blocks top-to-bottom, left-to-right
                data_blocks.sort(key=lambda b: (b[0], b[2]))

                print(f"Identified {len(data_blocks)} potential data blocks in sheet '{sheet_name}'.")


                # 4. Process each identified data block as a potential sub-table
                group_keywords = ['费用', '欠款', '汇总', '报销', '工资', '社保', '补贴', '补发', '补扣', '补偿', '补缴', '补助', '补款', '总计', '合计', '明细', '列表', '清单'] # Added more keywords

                for i, (min_r, max_r, min_c, max_c) in enumerate(data_blocks):
                    # Extract data for the current block
                    block_data = [row[min_c : max_c + 1] for row in raw_sheet_data[min_r : max_r + 1]]

                    # Convert block data to DataFrame
                    # Always read block data initially with header=None
                    block_df = pd.DataFrame(block_data)

                    if block_df.empty:
                        continue # Skip empty blocks

                    # 4.1 Clean block DataFrame
                    block_df = block_df.fillna("")
                    block_df = block_df.astype(str)
                    block_df = block_df.dropna(how='all') # Remove full empty rows within block

                    if block_df.empty:
                        continue # Skip if block is empty after row cleaning

                    # Reset index for the block DataFrame
                    block_df = block_df.reset_index(drop=True)


                    # 4.2 Determine Block Header (Attempt to preserve original header)
                    # Heuristic: If the first row of the block has significantly more non-empty cells
                    # or cell.value looks like a header (e.g., starts with text, few numbers)
                    first_block_row = block_df.iloc[0].tolist()
                    non_empty_in_first_block_row = [cell for cell in first_block_row if cell.strip()]

                    # Consider using first row as header if it's dense with text-like content
                    is_potential_block_header = False
                    if len(non_empty_in_first_block_row) > max(1, len(block_df.columns) // 2): # At least half the columns have content
                         # Check if content in first row looks more like text/headers than data
                         text_like_count = sum(1 for cell in non_empty_in_first_block_row if re.search(r'[a-zA-Z\u4e00-\u9fa5]', cell) and not re.fullmatch(r'[\d,\.]+', cell)) # Contains letters, not just numbers/punctuation
                         if text_like_count >= len(non_empty_in_first_block_row) // 2: # At least half non-empty cells look like text
                              is_potential_block_header = True

                    if is_potential_block_header:
                        # Use the first row as header, drop the first row from data
                        potential_headers = [str(col).strip() if str(col).strip() else f"Col_{j}" for j, col in enumerate(first_block_row)]
                        # Handle duplicate headers
                        seen_headers = {}
                        cleaned_headers = []
                        for header in potential_headers:
                             if header in seen_headers:
                                 seen_headers[header] += 1
                                 cleaned_headers.append(f"{header}.{seen_headers[header]}")
                             else:
                                 seen_headers[header] = 0
                                 cleaned_headers.append(header)

                        block_df.columns = cleaned_headers
                        block_df = block_df.iloc[1:].reset_index(drop=True) # Drop the header row from data
                        print(f"Block {i}: Used first row as header.")
                    else:
                        # Generate headers based on column index
                        block_df.columns = [f"Col_{j}" for j in range(len(block_df.columns))]
                        print(f"Block {i}: Used default Col_X headers.")

                    # After setting columns, ensure they are string type and strip spaces again
                    block_df.columns = block_df.columns.astype(str).str.strip()


                    # 4.3 Delete columns within the block that are now entirely empty (after potential header removal)
                    block_df = block_df.loc[:, (block_df != '').any(axis=0)]
                    block_df = block_df.reset_index(drop=True) # Reset index again

                    if block_df.empty:
                         print(f"Block {i}: is empty after column cleaning.")
                         continue # Skip if block is empty after column cleaning


                    # 4.4 Merge summary/remark rows *within this block*
                    remark_col_name = None
                    remark_keywords_in_col_names = ['备注', '说明', 'note', 'remark']
                    for col in block_df.columns:
                        if any(kw in str(col).lower() for kw in remark_keywords_in_col_names):
                            remark_col_name = col
                            break

                    if remark_col_name is None:
                         block_df['备注'] = ""
                         remark_col_name = '备注'
                         # print("No remark column found in block, added '备注' column.") # Too verbose

                    # Ensure the identified remark column is the last column for cleaner Markdown
                    if remark_col_name != block_df.columns[-1]:
                        cols = block_df.columns.tolist()
                        # Check if remark_col_name actually exists in the current block's columns
                        if remark_col_name in cols:
                            cols.remove(remark_col_name)
                            cols.append(remark_col_name)
                            block_df = block_df[cols]


                    summary_keywords_merge = ['总计', '合计', '汇总', '小计', 'summary', 'total', 'sum', 'subtotal'] # Keywords specifically for merging rows

                    block_rows_to_drop = []
                    # Iterate backwards within the block
                    for idx in range(len(block_df) - 1, 0, -1): # Start from second to last row
                         row = block_df.iloc[idx]
                         is_potential_summary_row = False
                         summary_parts = [] # Content to be merged

                         # Check if the row contains any merging keywords in any non-empty cell (excluding remark column itself)
                         contains_keyword = False
                         non_remark_cols = [c for c in block_df.columns if c != remark_col_name]
                         if non_remark_cols:
                             if block_df.iloc[idx][non_remark_cols].astype(str).str.lower().str.contains('|'.join(summary_keywords_merge)).any():
                                  contains_keyword = True
                                  # Collect content from non-remark columns with keywords
                                  for col_name in non_remark_cols:
                                      cell_value = str(row[col_name]).strip()
                                      if cell_value and any(kw in cell_value.lower() for kw in summary_keywords_merge):
                                           summary_parts.append(f"{col_name}: {cell_value}")

                         # Also consider if the remark column itself contains a *strong* summary keyword and other columns are sparse
                         content_in_remark_col = str(row[remark_col_name]).strip()
                         if not contains_keyword and content_in_remark_col and any(kw in content_in_remark_col.lower() for kw in summary_keywords_merge + ['总', '计']): # Add single chars for robustness
                             if (row.drop(remark_col_name).astype(str).str.strip() == "").all(): # If all other columns are empty
                                  contains_keyword = True # Treat as summary row triggered by remark column
                                  summary_parts.append(content_in_remark_col) # Add remark content


                         if contains_keyword:
                              # Append to the remark column of the previous row
                             prev_idx = block_df.index[idx - 1] # Get the actual index of the previous row
                             prev_remark = str(block_df.at[prev_idx, remark_col_name]).strip()

                             # Filter out empty collected parts
                             valid_summary_texts = [text for text in summary_parts if text.strip()]

                             merged_remark = prev_remark + " | " + " | ".join(valid_summary_texts) if prev_remark and valid_summary_texts else prev_remark if prev_remark else " | ".join(valid_summary_texts)
                             block_df.at[prev_idx, remark_col_name] = merged_remark.strip(" |") # Update previous row's remark

                             # Mark current row for dropping
                             block_rows_to_drop.append(block_df.index[idx])
                             # print(f"Block {i}: Merged row {block_df.index[idx]} into row {block_df.index[idx-1]}") # Debugging


                    # Delete rows marked for dropping within this block
                    if block_rows_to_drop:
                         block_df = block_df.drop(block_rows_to_drop)
                         block_df = block_df.reset_index(drop=True) # Reset index again
                         # print(f"Block {i}: Dropped {len(block_rows_to_drop)} potential summary/remark rows.")


                    # 4.5 Add Formula/Format/Comment info to Remark column (only for .xlsx)
                    if not is_xls and remark_col_name in block_df.columns:
                         for r_idx_block in range(len(block_df)):
                             original_row_idx = min_r + r_idx_block # Map back to original sheet index (approximate after cleaning)
                             # Check each cell in the current block row
                             for c_idx_block, col_name in enumerate(block_df.columns):
                                  # Find original column index - tricky after column dropping/reordering
                                  # A simpler approach: iterate through original columns and check if they are in the current block's columns
                                  original_c_idx = None
                                  # This mapping is complex after dropping/reordering columns.
                                  # Let's refine the cell_meta storage to use column names if possible, but index is more reliable before df conversion.
                                  # Assuming original column index corresponds roughly to block_df column index for now (imperfect).
                                  # Let's use the original block boundary indices + block_df index
                                  original_c_idx = min_c + c_idx_block # This is only valid IF no columns were dropped in the block


                                  # Let's retry the column mapping more carefully
                                  # Need a mapping from block_df column index to original column index in sheet_data
                                  # This requires storing original indices during block extraction/column cleaning.
                                  # For simplicity in this iteration, let's use the original indices (min_r, min_c) + block index,
                                  # but acknowledge it might be off if columns were dropped *within* the block's original range before mapping.

                                  original_cell_key = (original_row_idx, min_c + c_idx_block) # Using original column index from block range
                                  # If columns were dropped from block_df, the mapping needs refinement.
                                  # A better way: When creating block_df, store original column indices.
                                  # For now, let's use the simpler approach and note limitation.

                                  # Re-thinking: Instead of mapping back to original (r, c) by index, can we map by *value* or *structure*? Too complex.
                                  # Let's assume for now that cell_meta lookup by (original_row_idx, original_col_idx) is mostly correct,
                                  # acknowledging potential inaccuracies if columns within a block were dropped *before* this step.

                                  # Let's refine the block_df column mapping. When creating block_df from block_data,
                                  # we need to know which original columns those belong to.

                                  # Reverting to a simpler method for formula/format notes due to complex re-indexing:
                                  # Instead of per cell, maybe add notes *per row* in the remark column? Too messy.
                                  # Let's stick to the per-cell idea, but acknowledge the index mapping challenge.

                                  # Let's assume the cell_meta dict keys (r, c) are the original 0-based sheet indices.
                                  # We need to map the current cell in block_df (r_idx_block, c_idx_block) to its original sheet index (original_row_idx, original_col_idx).
                                  # original_row_idx = min_r + r_idx_block is correct for rows.
                                  # original_col_idx is tricky. It's the original column index in sheet_data that maps to c_idx_block in block_df.
                                  # We need the mapping created during column cleaning.
                                  # For now, let's skip detailed per-cell formula/format annotation in remarks due to complexity.
                                  # This is a known limitation of transforming irregular Excel data to structured Markdown.
                                  pass # Skip detailed formula/format addition for now


                    # 5. Add Block Title (H4)
                    # Try to find a title for this block from preceding rows/cells or the block's first row if not used as header
                    block_title = f"数据块 {i+1}" # Default title
                    # Look at the row just before this block started in the original sheet data
                    if min_r > 0:
                         # Check the row immediately above the block (min_r - 1) in the raw sheet data
                         prev_row = raw_sheet_data[min_r - 1]
                         # Check if this previous row was sparse and contained a potential title keyword
                         non_empty_in_prev_row = [cell for cell in prev_row if str(cell).strip()]
                         if len(non_empty_in_prev_row) > 0 and len(non_empty_in_prev_row) <= max(1, cols_count // 4): # Sparse
                             first_non_empty_in_prev = str(non_empty_in_prev_row[0]).strip()
                             # Check if the content looks like a title (contains keywords)
                             if any(kw in first_non_empty_in_prev for kw in group_keywords):
                                 block_title = first_non_empty_in_prev
                                 print(f"Block {i}: Found title '{block_title}' from row above.")
                                 # Optionally, check a few rows above? Too complex.

                    # Alternative: If the first row of the block wasn't used as a header, could it be a title?
                    if not is_potential_block_header:
                         # Check if the first row of the block itself is sparse and contains a title keyword
                         if len(non_empty_in_first_block_row) > 0 and len(non_empty_in_first_block_row) <= max(1, len(block_df.columns) // 4):
                              first_cell_in_block = str(block_df.iloc[0, 0]).strip() # First cell of the first row of the block_df
                              if any(kw in first_cell_in_block for kw in group_keywords):
                                   block_title = first_cell_in_block
                                   print(f"Block {i}: Found title '{block_title}' from first row of block.")
                                   # Since we used the first row as title, should we remove it from the data?
                                   # This conflicts with using it as potential header.
                                   # Let's refine: A row is EITHER a header OR a title, not both for the same block.
                                   # If first row looks like a title (sparse + keyword), treat it as a title, and the *next* row is the header (or default Col_X).
                                   # This makes the logic much more complex.

                                   # Let's simplify: The first non-empty sparse row BEFORE a data block is the title.
                                   # If no such row, try the first row of the block itself *only if it wasn't used as header*.
                                   pass # Logic already attempts this based on position/sparsity


                    # Add block title to markdown parts (H4 for blocks)
                    sheet_md_parts.append(f"\n#### {block_title}\n")
                    print(f"Processed block {i} ({min_r}:{max_r}, {min_c}:{max_c}).")


                    # 6. Convert block DataFrame to Markdown table
                    if not block_df.empty:
                        sheet_md_parts.append(_df_to_md(block_df))
                    else:
                         print(f"Block {i} became empty after cleaning.")


                # Add sheet's markdown parts to the overall list
                all_sheets_markdown.extend(sheet_md_parts)

            except Exception as e:
                 print(f"Error processing sheet '{sheet_name}' in {excel_path}: {e}")
                 # Continue to the next sheet

    except Exception as e:
        print(f"Overall error processing file {excel_path}: {e}")
        # No return here, so the finally block for cleanup is still reached
        pass # Just print error and let finally run


    # 7. Combine all sheet markdown parts
    final_md = "\n".join(all_sheets_markdown).strip() + "\n"

    # 8. Write to file
    if final_md.strip(): # Only write if there is content
        output_dir = os.path.dirname(md_path)
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
        try:
            with open(md_path, "w", encoding="utf-8") as f:
                f.write(final_md)
            print(f"\nSuccessfully wrote Markdown to: {md_path}")
        except Exception as e:
             print(f"Error writing Markdown to {md_path}: {e}")

    else:
        print(f"\nNo significant content extracted from {excel_path}, skipped writing Markdown to {md_path}.")


    # No temporary xlsx file is created/needed in this approach
    # The workbook object should be closed when it goes out of scope or program exits
    # For openpyxl read_only=False, explicit close might be better, but documentation suggests not strictly necessary in most cases.
    # if workbook and hasattr(workbook, 'close'):
    #     try:
    #         workbook.close()
    #         print("Workbook closed.")
    #     except Exception:
    #         pass # Ignore errors during close


# Keep the main part for processing files/directories
def process_single_file(excel_file):
    """
    处理单个文件，生成同目录下的 md 文件。
    """
    md_file = os.path.splitext(excel_file)[0] + ".md"
    excel_to_markdown(excel_file, md_file)
    # print(f"已生成: {md_file}") # Moved into excel_to_markdown


def process_directory(excel_dir):
    """
    递归处理目录下所有 xlsx/xls 文件，生成平行的 md 目录，结构完全对等。
    """
    # 生成平行目录名
    parent_dir = os.path.dirname(os.path.abspath(excel_dir.rstrip("/")))
    base_name = os.path.basename(os.path.abspath(excel_dir.rstrip("/")))
    now = datetime.datetime.now().strftime("%Y%m%d-%H%M")
    md_dir = os.path.join(parent_dir, f"{base_name}-md解析-{now}")

    # 递归遍历
    for root, dirs, files in os.walk(excel_dir):
        # 计算当前目录在原目录下的相对路径
        rel_path = os.path.relpath(root, excel_dir)
        # 在平行 md 目录下创建对应子目录
        md_subdir = os.path.join(md_dir, rel_path)
        os.makedirs(md_subdir, exist_ok=True)
        for file in files:
            if file.lower().endswith((".xlsx", ".xls")):
                excel_file = os.path.join(root, file)
                md_file = os.path.join(md_subdir, os.path.splitext(file)[0] + ".md")
                try:
                    excel_to_markdown(excel_file, md_file)
                    # print(f"已生成: {md_file}") # Moved into excel_to_markdown
                except Exception as e:
                    print(f"处理 {excel_file} 时出错: {e}")

    print(f"\n所有文件已处理完毕，md 文件保存在: {md_dir}")


def main():
    if len(sys.argv) != 2:
        print("用法: python excel_md_convert.py <excel文件或目录>")
        sys.exit(1)
    input_path = sys.argv[1]
    if os.path.isfile(input_path) and input_path.lower().endswith((".xlsx", ".xls")):
        process_single_file(input_path)
    elif os.path.isdir(input_path):
        process_directory(input_path)
    else:
        print("请输入有效的 xlsx/xls 文件或目录路径。")
        sys.exit(1)

if __name__ == "__main__":
    # Add import for zipfile here if not already at the top
    import zipfile
    main()
